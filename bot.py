#!/usr/bin/env python3
"""
Telegram-бот «Большой день женского здоровья» с Еленой Пшинник
Хранение данных: SQLite (локальный файл bot.db)
"""

import asyncio
import logging
import io
import os
import json
import sqlite3
from datetime import datetime, timedelta, timezone
from contextlib import contextmanager

from aiohttp import web

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, ChatJoinRequestHandler, ChatMemberHandler,
    filters, ContextTypes
)
from telegram.constants import ParseMode

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ═══════════════════════════════════════════════════════════
#  НАСТРОЙКИ — берутся из переменных окружения (Railway → Variables).
#  Значения справа от or — локальный fallback для запуска на своём ПК.
#
#  ⚠️ ВАЖНО: токен ниже когда-то попал в публичный репозиторий —
#  отзови его у @BotFather (/revoke) и задай новый через переменную BOT_TOKEN.
# ═══════════════════════════════════════════════════════════
def _env_admin_ids(name: str, default: list[int]) -> list[int]:
    raw = os.getenv(name, "")
    ids = [int(x) for x in raw.replace(" ", "").split(",") if x.strip().lstrip("-").isdigit()]
    return ids or default

BOT_TOKEN          = os.getenv("BOT_TOKEN", "8757283175:AAEy1joRPQl-QfFJ84QvtgqW1vgXaormSjg")
ADMIN_IDS          = _env_admin_ids("ADMIN_IDS", [334618540])

WORKSHOP_DATE_STR  = os.getenv("WORKSHOP_DATE_STR", "07 июня 2025")   # дата для показа
WORKSHOP_DATETIME  = datetime(2025, 7, 15, 10, 0)
WORKSHOP_PRICE     = int(os.getenv("WORKSHOP_PRICE", "10"))
WORKSHOP_LOCATION  = os.getenv("WORKSHOP_LOCATION", "Ссылка появится за день до воркшопа")
PAYMENT_PAGE_URL   = os.getenv("PAYMENT_PAGE_URL", "https://dr.pshinnik.ru/workshop_1")

QR_IMAGE_PATH      = os.getenv("QR_IMAGE_PATH", "qr_sbp.png")
DB_PATH            = os.getenv("DB_PATH", "bot.db")

# ── Приём вебхука от Тильды и автоматическая выдача доступа ──
# Секрет, который Тильда передаёт в URL: /tilda-webhook?secret=...  (защита эндпоинта)
TILDA_WEBHOOK_SECRET = os.getenv("TILDA_WEBHOOK_SECRET", "")
# ID закрытого Telegram-канала/группы (вида -1001234567890). Бот должен быть там админом
# с правом «Приглашать пользователей / создавать ссылки-приглашения».
_chan = os.getenv("WORKSHOP_CHANNEL_ID", "").strip()
WORKSHOP_CHANNEL_ID  = int(_chan) if _chan.lstrip("-").isdigit() else None
# Порт HTTP-сервера (Railway подставляет автоматически).
PORT                 = int(os.getenv("PORT", "8080"))

# Реквизиты ИП
IP_FIO             = "Пшинник Елена Борисовна"
IP_INN             = "622601705505"             # ← вставь ИНН
IP_OGRNIP          = "1027739609391"          # ← вставь ОГРНИП
IP_EMAIL           = "elena-pshinnik@mail.ru"
IP_ADDRESS         = "107031, Г. МОСКВА, г Москва, г МОСКВА, УЛ РОЖДЕСТВЕНКА, 10/2, СТР 1"

# ═══════════════════════════════════════════════════════════
#  СОСТОЯНИЯ
# ═══════════════════════════════════════════════════════════
(
    S_CONSENT, S_FULLNAME, S_PHONE, S_EMAIL,
    S_PAYMENT, S_SCREENSHOT,
    S_SUPPORT,
    S_BROADCAST_TARGET, S_BROADCAST_TEXT,
) = range(9)

# ═══════════════════════════════════════════════════════════
#  БАЗА ДАННЫХ (SQLite)
# ═══════════════════════════════════════════════════════════
@contextmanager
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    with db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS participants (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     INTEGER UNIQUE,
                username    TEXT,
                full_name   TEXT,
                phone       TEXT,
                email       TEXT,
                status      TEXT DEFAULT 'Ожидает подтверждения',
                amount      TEXT DEFAULT '',
                reg_date    TEXT,
                pay_date    TEXT DEFAULT '',
                note        TEXT DEFAULT ''
            )
        """)
        # Оплаты с сайта, для которых ещё нет участника в боте
        # (человек оплатил раньше, чем зарегистрировался). Матчим по email/телефону.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS pending_payments (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                email     TEXT DEFAULT '',
                phone     TEXT DEFAULT '',
                amount    TEXT DEFAULT '',
                created   TEXT
            )
        """)

def norm_phone(raw) -> str:
    """Приводит телефон к каноническому виду 7XXXXXXXXXX для сравнения."""
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if len(digits) == 11 and digits[0] == "8":
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    return digits[-11:] if len(digits) >= 11 else digits

def phone_tail10(raw) -> str:
    """Хвост из 10 цифр для мягкого сравнения номеров между системами."""
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    return digits[-10:] if len(digits) >= 10 else digits

def db_add(data: dict) -> bool:
    try:
        with db() as conn:
            conn.execute("""
                INSERT INTO participants
                (user_id, username, full_name, phone, email, reg_date)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    username=excluded.username,
                    full_name=excluded.full_name,
                    phone=excluded.phone,
                    email=excluded.email
            """, (
                data["user_id"], data.get("username", ""),
                data["full_name"], data["phone"], data["email"],
                datetime.now().strftime("%d.%m.%Y %H:%M")
            ))
        return True
    except Exception as e:
        logging.error(f"DB add error: {e}"); return False

def db_confirm(user_id: int, amount: str) -> bool:
    try:
        with db() as conn:
            conn.execute("""
                UPDATE participants
                SET status=?, amount=?, pay_date=?
                WHERE user_id=?
            """, ("Оплачено", amount, datetime.now().strftime("%d.%m.%Y %H:%M"), user_id))
        return True
    except Exception as e:
        logging.error(f"DB confirm error: {e}"); return False

def db_all() -> list:
    with db() as conn:
        rows = conn.execute("SELECT * FROM participants ORDER BY id").fetchall()
        return [dict(r) for r in rows]

def db_get(user_id: int) -> dict | None:
    with db() as conn:
        row = conn.execute("SELECT * FROM participants WHERE user_id=?", (user_id,)).fetchone()
        return dict(row) if row else None

def db_get_by_email(email: str) -> dict | None:
    """Поиск участника по email (без учёта регистра) — для матчинга оплаты с Тильды.
    Берём самую свежую запись на случай нескольких регистраций с одним email."""
    if not email:
        return None
    with db() as conn:
        row = conn.execute(
            "SELECT * FROM participants WHERE LOWER(email)=LOWER(?) ORDER BY id DESC LIMIT 1",
            (email.strip(),)
        ).fetchone()
        return dict(row) if row else None

def db_find_by_contact(email: str, phone: str) -> dict | None:
    """Ищет участника по email ИЛИ по нормализованному телефону."""
    rec = db_get_by_email(email) if email else None
    if rec:
        return rec
    target = norm_phone(phone)
    target10 = phone_tail10(phone)
    if target:
        with db() as conn:
            rows = conn.execute("SELECT * FROM participants ORDER BY id DESC").fetchall()
        for r in rows:
            row_phone = r["phone"]
            if norm_phone(row_phone) == target:
                return dict(r)
            if target10 and phone_tail10(row_phone) == target10:
                return dict(r)
    return None

def db_add_pending(email: str, phone: str, amount: str):
    """Сохраняет оплату с сайта, под которую ещё нет участника в боте.
    Дедуп: повторные вебхуки по тому же email/телефону не плодят строки."""
    e = (email or "").strip().lower()
    p = norm_phone(phone)
    with db() as conn:
        if e:
            conn.execute("DELETE FROM pending_payments WHERE email=?", (e,))
        if p:
            conn.execute("DELETE FROM pending_payments WHERE phone=?", (p,))
        conn.execute(
            "INSERT INTO pending_payments (email, phone, amount, created) VALUES (?, ?, ?, ?)",
            (e, p, str(amount), datetime.now().strftime("%d.%m.%Y %H:%M"))
        )

def db_pop_pending(email: str, phone: str) -> dict | None:
    """Находит ожидающую оплату по email/телефону, удаляет её и возвращает.
    Вызывается, когда человек регистрируется в боте после оплаты на сайте."""
    e = (email or "").strip().lower()
    p = norm_phone(phone)
    p10 = phone_tail10(phone)
    with db() as conn:
        row = None
        if e:
            row = conn.execute(
                "SELECT * FROM pending_payments WHERE email=? ORDER BY id DESC LIMIT 1", (e,)
            ).fetchone()
        if not row and p:
            row = conn.execute(
                "SELECT * FROM pending_payments WHERE phone=? ORDER BY id DESC LIMIT 1", (p,)
            ).fetchone()
        if not row and p10:
            rows = conn.execute("SELECT * FROM pending_payments ORDER BY id DESC").fetchall()
            for r in rows:
                if phone_tail10(r["phone"]) == p10:
                    row = r
                    break
        if row:
            conn.execute("DELETE FROM pending_payments WHERE id=?", (row["id"],))
            return dict(row)
    return None

def db_paid_ids() -> list[int]:
    with db() as conn:
        rows = conn.execute("SELECT user_id FROM participants WHERE status='Оплачено'").fetchall()
        return [r["user_id"] for r in rows]

def db_all_ids() -> list[int]:
    with db() as conn:
        rows = conn.execute("SELECT user_id FROM participants").fetchall()
        return [r["user_id"] for r in rows]

def db_stats() -> dict:
    with db() as conn:
        total  = conn.execute("SELECT COUNT(*) FROM participants").fetchone()[0]
        paid   = conn.execute("SELECT COUNT(*) FROM participants WHERE status='Оплачено'").fetchone()[0]
        amount = conn.execute("SELECT SUM(CAST(amount AS INTEGER)) FROM participants WHERE status='Оплачено'").fetchone()[0] or 0
        return {"total": total, "paid": paid, "pending": total - paid, "amount": amount}

# ═══════════════════════════════════════════════════════════
#  EXCEL ЭКСПОРТ
# ═══════════════════════════════════════════════════════════
def make_excel(records: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Участники"

    # Заголовки
    headers = ["№", "Дата регистрации", "ФИО", "Телефон", "Email",
               "Статус", "Сумма, ₽", "Дата оплаты", "Telegram ID", "Username"]
    header_fill = PatternFill("solid", fgColor="C6EFCE")
    header_font = Font(bold=True, color="276221")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Данные
    paid_fill   = PatternFill("solid", fgColor="EBF5E6")
    pend_fill   = PatternFill("solid", fgColor="FFF9E6")

    for i, r in enumerate(records, 1):
        is_paid = r.get("status") == "Оплачено"
        row_fill = paid_fill if is_paid else pend_fill
        vals = [
            i,
            r.get("reg_date", ""),
            r.get("full_name", ""),
            r.get("phone", ""),
            r.get("email", ""),
            r.get("status", ""),
            r.get("amount", ""),
            r.get("pay_date", ""),
            str(r.get("user_id", "")),
            r.get("username", ""),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # Ширина столбцов
    widths = [5, 18, 30, 18, 28, 22, 10, 18, 14, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w

    # Итоговая строка
    last = len(records) + 2
    stats = db_stats()
    ws.cell(last, 1, "Итого:")
    ws.cell(last, 1).font = Font(bold=True)
    ws.cell(last, 6, f"Оплатили: {stats['paid']} / {stats['total']}")
    ws.cell(last, 7, stats["amount"])
    ws.cell(last, 7).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ═══════════════════════════════════════════════════════════
#  ТЕКСТЫ
# ═══════════════════════════════════════════════════════════
T_WELCOME = """
🌸 *Добро пожаловать!*

Вы регистрируетесь на воркшоп
*«Большой день женского здоровья»*
с Еленой Пшинник

📅 Дата: *{date}*
💰 Стоимость: *{price} ₽*

Перед регистрацией необходимо ознакомиться с документами.
""".strip()

T_DOCS = """
📄 *Документы*

Ознакомьтесь перед регистрацией:

Нажимая *«✅ Согласна»*, вы подтверждаете:
— согласие с публичной офертой
— согласие на обработку персональных данных (ФЗ №152-ФЗ)
""".strip()

T_FULLNAME = "✏️ *Введите ваше ФИО*\n\nФамилия Имя Отчество — полностью.\n_Пример: Иванова Мария Сергеевна_"
T_PHONE    = "📱 *Введите номер телефона*\n\nФормат: +7XXXXXXXXXX\n\n_Или нажмите кнопку ниже_"
T_EMAIL    = "📧 *Шаг 2 из 2: введите электронную почту*\n\n_Пример: name@mail.ru_"

T_PAYMENT = """
💳 *Оплата участия*

Стоимость: *{price} ₽*

Оплатите участие на сайте по кнопке ниже.

После успешной оплаты бот автоматически подтвердит платёж и пришлёт персональную ссылку в закрытый Telegram-канал.
""".strip()

T_SCREENSHOT = "📸 *Пришлите скриншот оплаты*\n\nСделайте скриншот успешного платежа и отправьте сюда.\n\n_На скриншоте должны быть видны: сумма, дата и статус «Успешно»_"

T_SUCCESS = """
✅ *Заявка принята!*

Спасибо, {first_name}! 🌸

Скриншот оплаты получен и передан Елене Борисовне на проверку.
После подтверждения вы получите доступ в закрытый Telegram-канал.

📋 *Ваши данные:*
👤 {full_name}
📱 {phone}
📧 {email}

📅 Воркшоп: *{date}*

🔗 *Ссылку на Telegram-канал воркшопа* мы пришлём за день до мероприятия — следите за сообщениями от бота!

_Вопросы — кнопка «💬 Поддержка» в меню._
""".strip()

T_PAYMENT_CONFIRMED = """
✅ *Оплата подтверждена!*

Спасибо, {full_name}! Ваше место на воркшопе забронировано 🌸

Елена Борисовна получила информацию о вашей оплате и рада видеть вас на воркшопе!

🔗 *Ссылка в закрытый Telegram-канал* отправлена в этом чате.
Если ссылка потерялась, нажмите кнопку *«🔗 Ссылка в канал»* в меню.

📅 Воркшоп: *{date}*

_Если возникнут вопросы — кнопка «💬 Поддержка» в меню._
""".strip()

T_ADMIN_NEW = """
🔔 *Новая заявка!*

👤 {full_name}
📱 {phone}
📧 {email}
🕐 {time}
🆔 `{user_id}` @{username}

✅ Подтвердить оплату:
`/confirm_{user_id}_{price}`
""".strip()

T_REMINDER = """
🌸 *Напоминание!*

Завтра — воркшоп
*«Большой день женского здоровья»*
с Еленой Пшинник 💫

📅 *{date}*
📍 {location}

До встречи! 🙏
""".strip()

T_OFFER_FULL = """
📋 *ПУБЛИЧНАЯ ОФЕРТА*
━━━━━━━━━━━━━━━━━━━━

*Индивидуальный предприниматель Пшинник Елена Борисовна*
ИНН: {inn} | ОГРНИП: {ogrnip}
Адрес: {address}
E-mail: {email}

*1. ПРЕДМЕТ ОФЕРТЫ*
Организатор проводит онлайн-воркшоп «Большой день женского здоровья» и предоставляет Участнику доступ к нему.
Дата: {workshop_date}. Формат: онлайн-трансляция в Telegram.

*2. АКЦЕПТ*
Акцептом является заполнение регистрационной формы и оплата. Акцепт = полное согласие с офертой.

*3. СТОИМОСТЬ И ОПЛАТА*
Стоимость: {price} рублей. Оплата через СБП до начала воркшопа.

*4. ОБЯЗАТЕЛЬСТВА СТОРОН*
Организатор: провести воркшоп, предоставить ссылку, вернуть оплату при отмене по вине организатора.
Участник: предоставить достоверные данные, не записывать и не распространять материалы.

*5. ВОЗВРАТ*
Возврат при отказе за 3+ дня до воркшопа. При отказе менее чем за 3 дня — оплата не возвращается. При отмене организатором — полный возврат в течение 10 рабочих дней.

*6. ОТВЕТСТВЕННОСТЬ*
Организатор не несёт ответственности за технические сбои на стороне Участника.

*7. ПЕРСОНАЛЬНЫЕ ДАННЫЕ*
Участник даёт согласие на обработку персональных данных согласно Политике конфиденциальности.

ИП Пшинник Елена Борисовна
━━━━━━━━━━━━━━━━━━━━
""".strip()

T_PRIVACY_FULL = """
🔒 *ПОЛИТИКА ОБРАБОТКИ ПЕРСОНАЛЬНЫХ ДАННЫХ*
━━━━━━━━━━━━━━━━━━━━

*ИП Пшинник Елена Борисовна*
ИНН: {inn} | E-mail: {email}

*1. ОБЩИЕ ПОЛОЖЕНИЯ*
Политика разработана в соответствии с ФЗ №152-ФЗ «О персональных данных».

*2. ЦЕЛИ ОБРАБОТКИ*
— Идентификация участника при регистрации
— Направление информации о воркшопе
— Подтверждение оплаты и выдача доступа
— Направление напоминаний и организационных сообщений

*3. ОБРАБАТЫВАЕМЫЕ ДАННЫЕ*
— Фамилия, имя, отчество
— Номер телефона
— Адрес электронной почты
— Идентификатор Telegram

*4. ПРАВОВОЕ ОСНОВАНИЕ*
Согласие субъекта персональных данных (ст. 6 ч. 1 п. 1 ФЗ-152).

*5. ПОРЯДОК ОБРАБОТКИ*
Данные хранятся в защищённой базе данных на сервере. Передача третьим лицам без согласия не осуществляется.

*6. СРОК ХРАНЕНИЯ*
1 год с момента проведения воркшопа.

*7. ПРАВА СУБЪЕКТА*
Вы вправе запросить, уточнить или удалить свои данные. Обращайтесь: {email}

ИП Пшинник Елена Борисовна
━━━━━━━━━━━━━━━━━━━━
""".strip()

# ═══════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНОЕ
# ═══════════════════════════════════════════════════════════
def is_admin(uid): return uid in ADMIN_IDS

def main_kb(user_id: int | None = None):
    """Главное меню: для оплативших скрываем оплату, для неоплативших показываем."""
    rec = db_get(user_id) if user_id else None
    is_paid = rec and rec.get("status") == "Оплачено"
    rows = [["📋 Моя регистрация", "💬 Поддержка"]]
    if not is_paid:
        rows.append(["💳 Повторить оплату", "🔄 Проверить оплату"])
    else:
        rows.append(["🔗 Ссылка в канал"])
    rows.append(["✏️ Изменить данные", "📞 Связаться с организатором"])
    rows.append(["📄 Оферта", "🔒 Политика данных"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)

def is_paid_user(user_id: int) -> bool:
    rec = db_get(user_id)
    return bool(rec and rec.get("status") == "Оплачено")

async def send_channel_invite(bot, user_id: int, full_name: str) -> bool:
    """Отправляет персональную ссылку в закрытый канал.
    Ссылка создаётся как join-request: бот одобрит только оплаченных."""
    if not WORKSHOP_CHANNEL_ID:
        await bot.send_message(
            chat_id=user_id,
            text="⚠️ Канал пока не настроен. Напишите в поддержку, мы пришлём доступ вручную."
        )
        return False
    try:
        invite = await bot.create_chat_invite_link(
            chat_id=WORKSHOP_CHANNEL_ID,
            creates_join_request=True,
            expire_date=datetime.now(timezone.utc) + timedelta(days=7),
            name=f"{full_name[:28]}",
        )
        await bot.send_message(
            chat_id=user_id,
            text=("🔗 *Ваша ссылка в закрытый канал воркшопа:*\n\n"
                  f"{invite.invite_link}\n\n"
                  "После перехода нажмите «Join». Бот автоматически одобрит доступ, "
                  "если оплата подтверждена."),
            parse_mode=ParseMode.MARKDOWN,
        )
        return True
    except Exception as e:
        logging.error(f"send_channel_invite error for {user_id}: {e}")
        await bot.send_message(
            chat_id=user_id,
            text=("⚠️ Не удалось создать ссылку в канал автоматически.\n"
                  "Напишите в поддержку, мы выдадим доступ вручную."),
        )
        for aid in ADMIN_IDS:
            try:
                await bot.send_message(
                    chat_id=aid,
                    text=(f"⚠️ Ошибка выдачи ссылки в канал: {e}\n"
                          f"user_id `{user_id}`, ФИО: {full_name}"),
                    parse_mode=ParseMode.MARKDOWN,
                )
            except Exception:
                pass
        return False

async def has_channel_access(bot, user_id: int) -> bool:
    """Проверяет, состоит ли пользователь в закрытом канале воркшопа."""
    if not WORKSHOP_CHANNEL_ID:
        return False
    try:
        member = await bot.get_chat_member(WORKSHOP_CHANNEL_ID, user_id)
        # left/kicked — доступа нет. Остальные статусы считаем валидным доступом.
        return member.status not in ("left", "kicked")
    except Exception as e:
        logging.warning(f"Channel access check failed for {user_id}: {e}")
        return False

# ═══════════════════════════════════════════════════════════
#  /start
# ═══════════════════════════════════════════════════════════
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    arg = ctx.args[0] if ctx.args else ""
    if arg == "offer":   return await _send_offer(update)
    if arg == "privacy": return await _send_privacy(update)
    # Если пользователь уже зарегистрирован — не запускаем анкету заново.
    rec = db_get(update.effective_user.id)
    if rec:
        if rec["status"] == "Оплачено":
            await update.message.reply_text(
                f"🌸 Вы уже зарегистрированы и оплата подтверждена.\n\n"
                f"📅 Воркшоп: {WORKSHOP_DATE_STR}\n"
                f"👤 {rec['full_name']}",
                reply_markup=main_kb(update.effective_user.id)
            )
        else:
            kb = [[InlineKeyboardButton("💳 Оплатить на сайте", url=PAYMENT_PAGE_URL)]]
            await update.message.reply_text(
                f"Вы уже зарегистрированы, но оплата пока не подтверждена.\n\n"
                f"💰 Стоимость: {WORKSHOP_PRICE} ₽\n"
                f"Нажмите кнопку, чтобы перейти к оплате:",
                reply_markup=InlineKeyboardMarkup(kb)
            )
            await update.message.reply_text(
                "После оплаты нажмите «🔄 Проверить оплату» в меню.",
                reply_markup=main_kb(update.effective_user.id)
            )
        return ConversationHandler.END
    # Запись могла потеряться (например, после рестарта), но доступ уже выдан.
    # Если пользователь уже состоит в закрытом канале — восстанавливаем статус.
    if await has_channel_access(ctx.bot, update.effective_user.id):
        user = update.effective_user
        db_add({
            "user_id": user.id,
            "username": user.username or "",
            "full_name": user.full_name or "Участник",
            "phone": "",
            "email": "",
        })
        db_confirm(user.id, str(WORKSHOP_PRICE))
        await update.message.reply_text(
            "✅ Доступ уже активен. Вижу вас среди участников закрытого канала.",
            reply_markup=main_kb(user.id),
        )
        return ConversationHandler.END
    ctx.user_data.clear()
    kb = [[InlineKeyboardButton("📄 Ознакомиться с документами →", callback_data="show_docs")]]
    await update.message.reply_text(
        T_WELCOME.format(date=WORKSHOP_DATE_STR, price=WORKSHOP_PRICE),
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb)
    )
    return S_CONSENT

async def _send_offer(update):
    text = T_OFFER_FULL.format(inn=IP_INN, ogrnip=IP_OGRNIP, address=IP_ADDRESS,
                                email=IP_EMAIL, workshop_date=WORKSHOP_DATE_STR, price=WORKSHOP_PRICE)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    return S_CONSENT

async def _send_privacy(update):
    text = T_PRIVACY_FULL.format(inn=IP_INN, email=IP_EMAIL)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    return S_CONSENT

# ═══════════════════════════════════════════════════════════
#  РЕГИСТРАЦИЯ
# ═══════════════════════════════════════════════════════════
async def cb_show_docs(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    kb = [
        [InlineKeyboardButton("📋 Публичная оферта",               callback_data="read_offer")],
        [InlineKeyboardButton("🔒 Политика персональных данных",   callback_data="read_privacy")],
        [InlineKeyboardButton("✅ Согласна со всеми условиями",    callback_data="consent_yes")],
        [InlineKeyboardButton("❌ Не согласна",                    callback_data="consent_no")],
    ]
    await q.edit_message_text(T_DOCS, parse_mode=ParseMode.MARKDOWN,
                              reply_markup=InlineKeyboardMarkup(kb))
    return S_CONSENT

async def cb_read_offer(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    text = T_OFFER_FULL.format(inn=IP_INN, ogrnip=IP_OGRNIP, address=IP_ADDRESS,
                                email=IP_EMAIL, workshop_date=WORKSHOP_DATE_STR, price=WORKSHOP_PRICE)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await q.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    kb = [[InlineKeyboardButton("✅ Согласна", callback_data="consent_yes"),
           InlineKeyboardButton("❌ Не согласна", callback_data="consent_no")]]
    await q.message.reply_text("Продолжить?", reply_markup=InlineKeyboardMarkup(kb))
    return S_CONSENT

async def cb_read_privacy(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    text = T_PRIVACY_FULL.format(inn=IP_INN, email=IP_EMAIL)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await q.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    kb = [[InlineKeyboardButton("✅ Согласна", callback_data="consent_yes"),
           InlineKeyboardButton("❌ Не согласна", callback_data="consent_no")]]
    await q.message.reply_text("Продолжить?", reply_markup=InlineKeyboardMarkup(kb))
    return S_CONSENT

async def cb_consent_yes(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    ctx.user_data["user_id"]  = q.from_user.id
    ctx.user_data["username"] = q.from_user.username or ""
    ctx.user_data["full_name"] = q.from_user.full_name or "Участник"
    await q.edit_message_text("✅ Согласие получено! Начинаем регистрацию.")
    kb = [[KeyboardButton("📱 Отправить мой номер", request_contact=True)]]
    await q.message.reply_text(
        "📱 *Шаг 1 из 2: номер телефона*\n\nОтправьте ваш номер или введите вручную в формате +7XXXXXXXXXX.",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True)
    )
    return S_PHONE

async def cb_consent_no(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    await q.edit_message_text("❌ Без согласия регистрация невозможна.\n\nЕсли передумаете — /start")
    return ConversationHandler.END

async def rx_fullname(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if len(text.split()) < 2:
        await update.message.reply_text("⚠️ Введите полное ФИО (минимум имя и фамилию).")
        return S_FULLNAME
    ctx.user_data["full_name"] = text
    kb = [[KeyboardButton("📱 Отправить мой номер", request_contact=True)]]
    await update.message.reply_text(
        T_PHONE, parse_mode=ParseMode.MARKDOWN,
        reply_markup=ReplyKeyboardMarkup(kb, resize_keyboard=True, one_time_keyboard=True)
    )
    return S_PHONE

async def rx_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.contact:
        phone = update.message.contact.phone_number
        if not phone.startswith("+"): phone = "+" + phone
    else:
        phone = update.message.text.strip()
        clean = phone.replace("+","").replace("-","").replace(" ","").replace("(","").replace(")","")
        if not clean.isdigit() or len(clean) < 10:
            await update.message.reply_text("⚠️ Введите корректный номер (+79991234567).")
            return S_PHONE
    ctx.user_data["phone"] = phone
    if not ctx.user_data.get("full_name"):
        ctx.user_data["full_name"] = update.effective_user.full_name or "Участник"
    await update.message.reply_text(T_EMAIL, parse_mode=ParseMode.MARKDOWN,
                                    reply_markup=ReplyKeyboardRemove())
    return S_EMAIL

async def rx_email(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    email = update.message.text.strip().lower()
    if "@" not in email or "." not in email.split("@")[-1]:
        await update.message.reply_text("⚠️ Введите корректный email.")
        return S_EMAIL
    data = ctx.user_data
    data["email"] = email
    # Если пользователь уже есть по email/телефону — подтягиваем сохранённое ФИО.
    rec_existing = db_find_by_contact(email, data.get("phone", ""))
    if rec_existing and rec_existing.get("full_name"):
        data["full_name"] = rec_existing["full_name"]
    # Если по контактам уже есть оплаченный участник — сразу восстанавливаем доступ.
    if rec_existing and rec_existing.get("status") == "Оплачено":
        db_add({"user_id": data["user_id"], "username": data.get("username", ""),
                "full_name": data["full_name"], "phone": data["phone"], "email": email})
        db_confirm(data["user_id"], rec_existing.get("amount") or str(WORKSHOP_PRICE))
        rec = db_get(data["user_id"])
        await update.message.reply_text(
            "✅ Оплата уже найдена по вашим данным. Восстанавливаю доступ 🌸",
            parse_mode=ParseMode.MARKDOWN, reply_markup=main_kb(data["user_id"])
        )
        try:
            await grant_access(ctx.bot, rec, rec_existing.get("amount") or str(WORKSHOP_PRICE))
        except Exception as e:
            logging.error(f"grant_access (existing paid) error: {e}")
        return ConversationHandler.END

    # Проверяем: вдруг человек уже оплатил на сайте до регистрации в боте?
    pend = db_pop_pending(email, data.get("phone", ""))
    if pend:
        db_add({"user_id": data["user_id"], "username": data.get("username", ""),
                "full_name": data["full_name"], "phone": data["phone"], "email": email})
        amount = pend.get("amount") or str(WORKSHOP_PRICE)
        db_confirm(data["user_id"], amount)
        rec = db_get(data["user_id"])
        await update.message.reply_text(
            "✅ *Мы нашли вашу оплату на сайте!*\nОткрываю доступ 🌸",
            parse_mode=ParseMode.MARKDOWN, reply_markup=main_kb(data["user_id"])
        )
        try:
            await grant_access(ctx.bot, rec, amount)
        except Exception as e:
            logging.error(f"grant_access (pending) error: {e}")
        for aid in ADMIN_IDS:
            try:
                await ctx.bot.send_message(
                    chat_id=aid,
                    text=(f"✅ Авто-доступ: участник зарегистрировался после оплаты на сайте.\n"
                          f"👤 {data['full_name']}\n📧 {email}  📱 {data['phone']}\n"
                          f"💰 {amount} ₽  🆔 `{data['user_id']}`"),
                    parse_mode=ParseMode.MARKDOWN,
                )
            except Exception:
                pass
        remind_at = WORKSHOP_DATETIME - timedelta(days=1)
        if remind_at > datetime.now():
            ctx.job_queue.run_once(job_reminder, when=remind_at,
                                   data={"user_id": data["user_id"]},
                                   name=f"reminder_{data['user_id']}")
        return ConversationHandler.END

    # Регистрируем участника сразу, чтобы вебхук с сайта мог сматчить оплату
    db_add({"user_id": data["user_id"], "username": data.get("username", ""),
            "full_name": data["full_name"], "phone": data["phone"], "email": email})

    kb = [[InlineKeyboardButton("💳 Оплатить на сайте", url=PAYMENT_PAGE_URL)]]
    await update.message.reply_text(
        T_PAYMENT.format(price=WORKSHOP_PRICE),
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb)
    )
    await update.message.reply_text(
        "После оплаты вернитесь в бот — подтверждение и доступ придут автоматически.",
        reply_markup=main_kb(data["user_id"])
    )
    return ConversationHandler.END

async def cb_payment_done(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    await q.message.reply_text(T_SCREENSHOT, parse_mode=ParseMode.MARKDOWN)
    return S_SCREENSHOT

async def rx_screenshot(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.photo and not update.message.document:
        await update.message.reply_text(
            "⚠️ Пришлите *скриншот* (фото из галереи).", parse_mode=ParseMode.MARKDOWN)
        return S_SCREENSHOT

    file_id = (update.message.photo[-1].file_id if update.message.photo
               else update.message.document.file_id)
    data = ctx.user_data
    user = update.message.from_user
    parts = data["full_name"].split()
    first = parts[1] if len(parts) > 1 else parts[0]

    db_add({"user_id": user.id, "username": data.get("username",""),
            "full_name": data["full_name"], "phone": data["phone"], "email": data["email"]})

    await update.message.reply_text(
        T_SUCCESS.format(first_name=first, full_name=data["full_name"],
                         phone=data["phone"], email=data["email"], date=WORKSHOP_DATE_STR),
        parse_mode=ParseMode.MARKDOWN, reply_markup=main_kb(user.id)
    )

    # Уведомляем администратора
    for aid in ADMIN_IDS:
        try:
            await ctx.bot.send_message(
                chat_id=aid,
                text=T_ADMIN_NEW.format(
                    full_name=data["full_name"], phone=data["phone"], email=data["email"],
                    time=datetime.now().strftime("%d.%m.%Y %H:%M"),
                    user_id=user.id, username=data.get("username","—"), price=WORKSHOP_PRICE
                ), parse_mode=ParseMode.MARKDOWN
            )
            caption = f"💳 Скриншот от {data['full_name']}"
            if update.message.photo:
                await ctx.bot.send_photo(chat_id=aid, photo=file_id, caption=caption)
            else:
                await ctx.bot.send_document(chat_id=aid, document=file_id, caption=caption)
        except Exception as e:
            logging.error(f"Admin notify {aid}: {e}")

    # Напоминание за день
    remind_at = WORKSHOP_DATETIME - timedelta(days=1)
    if remind_at > datetime.now():
        ctx.job_queue.run_once(
            job_reminder, when=remind_at,
            data={"user_id": user.id},
            name=f"reminder_{user.id}"
        )
    return ConversationHandler.END

async def job_reminder(ctx: ContextTypes.DEFAULT_TYPE):
    uid = ctx.job.data["user_id"]
    try:
        await ctx.bot.send_message(
            chat_id=uid,
            text=T_REMINDER.format(date=WORKSHOP_DATE_STR, location=WORKSHOP_LOCATION),
            parse_mode=ParseMode.MARKDOWN
        )
    except Exception as e:
        logging.error(f"Reminder {uid}: {e}")

# ═══════════════════════════════════════════════════════════
#  ПОДДЕРЖКА
# ═══════════════════════════════════════════════════════════
async def start_support(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    kb = [[InlineKeyboardButton("❌ Отмена", callback_data="support_cancel")]]
    await update.message.reply_text(
        "💬 *Напишите ваш вопрос* — отвечу в ближайшее время.",
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb)
    )
    return S_SUPPORT

async def rx_support(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    text = update.message.text or "[вложение]"
    for aid in ADMIN_IDS:
        try:
            await ctx.bot.send_message(
                chat_id=aid,
                text=f"💬 *Вопрос поддержки*\n\n"
                     f"👤 {user.full_name} (@{user.username or '—'})\n"
                     f"🆔 `{user.id}`\n\n"
                     f"{text}\n\n"
                     f"_Ответить: `/reply_{user.id} текст`_",
                parse_mode=ParseMode.MARKDOWN
            )
        except Exception as e:
            logging.error(f"Support fwd {aid}: {e}")
    await update.message.reply_text("✅ Вопрос отправлен! Отвечу в ближайшее время 🙏",
                                    reply_markup=main_kb(update.effective_user.id))
    return ConversationHandler.END

async def cb_support_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    await q.edit_message_text("Возврат в меню.")
    return ConversationHandler.END

# ═══════════════════════════════════════════════════════════
#  МЕНЮ
# ═══════════════════════════════════════════════════════════
async def menu_my_reg(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rec = db_get(update.effective_user.id)
    if not rec:
        await update.message.reply_text(
            "Вы ещё не зарегистрированы.\nНажмите /start чтобы начать.",
            reply_markup=main_kb(update.effective_user.id)); return
    e = "✅" if rec["status"] == "Оплачено" else "⏳"
    await update.message.reply_text(
        f"📋 *Ваша регистрация*\n\n"
        f"👤 {rec['full_name']}\n📱 {rec['phone']}\n📧 {rec['email']}\n\n"
        f"{e} Статус: *{rec['status']}*\n📅 {WORKSHOP_DATE_STR}",
        parse_mode=ParseMode.MARKDOWN, reply_markup=main_kb(update.effective_user.id)
    )

async def menu_contact(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка «📞 Связаться с организатором»"""
    await update.message.reply_text(
        "📞 *Связаться с организатором*\n\n"
        "Если у вас есть вопросы — напишите напрямую в Telegram:\n\n"
        "👉 @whatshappened\n\n"
        "_Или нажмите кнопку «💬 Поддержка» — ваш вопрос придёт организатору прямо в бот._",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=main_kb(update.effective_user.id)
    )

async def menu_offer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = T_OFFER_FULL.format(inn=IP_INN, ogrnip=IP_OGRNIP, address=IP_ADDRESS,
                                email=IP_EMAIL, workshop_date=WORKSHOP_DATE_STR, price=WORKSHOP_PRICE)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)

async def menu_privacy(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = T_PRIVACY_FULL.format(inn=IP_INN, email=IP_EMAIL)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)

async def menu_resend_qr(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка «💳 Повторить оплату» — отправить ссылку на оплату на сайте."""
    rec = db_get(update.effective_user.id)
    if not rec:
        await update.message.reply_text(
            "Вы ещё не зарегистрированы. Нажмите /start чтобы начать.",
            reply_markup=main_kb(update.effective_user.id)); return
    if rec["status"] == "Оплачено":
        await update.message.reply_text(
            "✅ Ваша оплата уже подтверждена! Ничего делать не нужно.",
            reply_markup=main_kb(update.effective_user.id)); return
    kb = [[InlineKeyboardButton("💳 Оплатить на сайте", url=PAYMENT_PAGE_URL)]]
    await update.message.reply_text(
        T_PAYMENT.format(price=WORKSHOP_PRICE),
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb)
    )
    await update.message.reply_text(
        "После оплаты подтверждение и доступ в канал придут автоматически."
    )

async def menu_edit_data(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    """Кнопка «✏️ Изменить данные» — ввести данные заново без повторного согласия."""
    rec = db_get(update.effective_user.id)
    if rec:
        ctx.user_data["user_id"]  = update.effective_user.id
        ctx.user_data["username"] = update.effective_user.username or ""
        # Подставляем текущие данные как подсказку
        await update.message.reply_text(
            f"✏️ *Изменение данных*\n\n"
            f"Текущие данные:\n"
            f"👤 {rec['full_name']}\n"
            f"📱 {rec['phone']}\n"
            f"📧 {rec['email']}\n\n"
            f"Введите новое ФИО (или отправьте то же самое если менять не нужно):",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=ReplyKeyboardRemove()
        )
        return S_FULLNAME
    else:
        await update.message.reply_text(
            "Вы ещё не зарегистрированы. Нажмите /start чтобы начать.",
            reply_markup=main_kb(update.effective_user.id))
        return ConversationHandler.END

async def menu_check_payment(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка «🔄 Проверить оплату» — не требует повторной регистрации."""
    uid = update.effective_user.id
    rec = db_get(uid)
    if not rec:
        await update.message.reply_text(
            "Вы ещё не зарегистрированы. Нажмите /start, заполните данные и оплатите на сайте.",
            reply_markup=main_kb(uid),
        )
        return

    if rec["status"] == "Оплачено":
        await update.message.reply_text(
            "✅ Оплата уже подтверждена. Доступ к воркшопу открыт.",
            reply_markup=main_kb(uid),
        )
        return

    kb = [[InlineKeyboardButton("💳 Оплатить на сайте", url=PAYMENT_PAGE_URL)]]
    await update.message.reply_text(
        "⏳ Пока не вижу подтверждения оплаты.\n"
        "Если вы только что оплатили, подождите 10-30 секунд и нажмите «🔄 Проверить оплату» ещё раз.",
        reply_markup=InlineKeyboardMarkup(kb),
    )

async def menu_send_channel_link(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка «🔗 Ссылка в канал» — повторная выдача ссылки оплаченному пользователю."""
    uid = update.effective_user.id
    rec = db_get(uid)
    if not rec:
        await update.message.reply_text(
            "Вы ещё не зарегистрированы. Нажмите /start.",
            reply_markup=main_kb(uid),
        )
        return
    if rec["status"] != "Оплачено":
        kb = [[InlineKeyboardButton("💳 Оплатить на сайте", url=PAYMENT_PAGE_URL)]]
        await update.message.reply_text(
            "Ссылка в канал доступна только после подтверждённой оплаты.",
            reply_markup=InlineKeyboardMarkup(kb),
        )
        return
    ok = await send_channel_invite(
        ctx.bot,
        uid,
        rec.get("full_name") or update.effective_user.full_name or "Участник"
    )
    if ok:
        await update.message.reply_text(
            "Если ссылка не открылась, нажмите кнопку ещё раз.",
            reply_markup=main_kb(uid)
        )
    else:
        await update.message.reply_text(
            "Проверьте сообщения выше. Если ссылка не пришла — нажмите «💬 Поддержка».",
            reply_markup=main_kb(uid)
        )

async def on_join_request(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Одобряет заявку в канал только для оплаченных пользователей."""
    req = update.chat_join_request
    if not req:
        return
    uid = req.from_user.id
    chat_id = req.chat.id
    if WORKSHOP_CHANNEL_ID and chat_id != WORKSHOP_CHANNEL_ID:
        return
    if is_paid_user(uid):
        await ctx.bot.approve_chat_join_request(chat_id=chat_id, user_id=uid)
        try:
            await ctx.bot.send_message(uid, "✅ Доступ в канал подтверждён. Добро пожаловать!")
        except Exception:
            pass
    else:
        await ctx.bot.decline_chat_join_request(chat_id=chat_id, user_id=uid)
        try:
            await ctx.bot.send_message(
                uid,
                f"❌ Доступ в канал только после оплаты.\nОплатить: {PAYMENT_PAGE_URL}"
            )
        except Exception:
            pass

async def on_chat_member_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Если неоплаченный пользователь уже вошёл в канал — удаляем его."""
    cm = update.chat_member
    if not cm:
        return
    if WORKSHOP_CHANNEL_ID and cm.chat.id != WORKSHOP_CHANNEL_ID:
        return
    uid = cm.new_chat_member.user.id
    new_status = cm.new_chat_member.status
    if new_status in ("member", "administrator", "creator") and not is_paid_user(uid):
        try:
            await ctx.bot.ban_chat_member(chat_id=cm.chat.id, user_id=uid)
            await ctx.bot.unban_chat_member(chat_id=cm.chat.id, user_id=uid, only_if_banned=True)
        except Exception as e:
            logging.warning(f"Failed to remove unpaid user {uid}: {e}")

async def cb_go_support(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    """Кнопка поддержки из меню оплаты."""
    q = update.callback_query; await q.answer()
    kb = [[InlineKeyboardButton("❌ Отмена", callback_data="support_cancel")]]
    await q.message.reply_text(
        "💬 *Напишите ваш вопрос* — отвечу в ближайшее время.",
        parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(kb)
    )
    return S_SUPPORT

# ═══════════════════════════════════════════════════════════
#  ADMIN
# ═══════════════════════════════════════════════════════════
async def cmd_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    try:
        _, uid_s, amount = update.message.text.split("_", 2)
        user_id = int(uid_s)
    except ValueError:
        await update.message.reply_text("Формат: `/confirm_123456789_10000`",
                                        parse_mode=ParseMode.MARKDOWN); return
    db_confirm(user_id, amount)
    rec = db_get(user_id)
    full_name = rec["full_name"] if rec else "Участник"
    if not rec:
        rec = {"user_id": user_id, "full_name": full_name}
    try:
        await grant_access(ctx.bot, rec, amount)
        await update.message.reply_text(f"✅ Подтверждено. Сообщение и ссылка отправлены — {full_name}.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Не удалось отправить сообщение: {e}")

async def cmd_reply(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    try:
        parts = update.message.text.split(" ", 1)
        user_id = int(parts[0].split("_")[1])
        text = parts[1] if len(parts) > 1 else ""
    except (IndexError, ValueError):
        await update.message.reply_text("Формат: `/reply_123456789 Текст ответа`",
                                        parse_mode=ParseMode.MARKDOWN); return
    try:
        await ctx.bot.send_message(chat_id=user_id,
                                   text=f"💬 *Ответ организатора:*\n\n{text}",
                                   parse_mode=ParseMode.MARKDOWN)
        await update.message.reply_text("✅ Ответ отправлен.")
    except Exception as e:
        await update.message.reply_text(f"⚠️ Ошибка: {e}")

async def cmd_participants(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа."); return
    records = db_all()
    if not records:
        await update.message.reply_text("📋 Участников пока нет."); return
    st = db_stats()
    lines = [f"📋 *Участники* — {st['total']} чел.\n"
             f"✅ Оплатили: {st['paid']} | ⏳ Ожидают: {st['pending']} | 💰 {st['amount']} ₽\n"
             f"{'━'*20}"]
    for r in records:
        e = "✅" if r["status"] == "Оплачено" else "⏳"
        amt = f" | {r['amount']} ₽" if r.get("amount") else ""
        lines.append(f"{e} *{r['id']}.* {r['full_name']}\n"
                     f"   📱 {r['phone']}  📧 {r['email']}{amt}")
    text = "\n".join(lines)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
    kb = [[InlineKeyboardButton("📥 Скачать Excel", callback_data="export_excel")]]
    await update.message.reply_text("Выгрузить в файл?", reply_markup=InlineKeyboardMarkup(kb))

async def cb_export_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    if not is_admin(q.from_user.id): return
    records = db_all()
    if not records:
        await q.message.reply_text("Нет данных."); return
    if not HAS_OPENPYXL:
        await q.message.reply_text("⚠️ Установи openpyxl: pip install openpyxl"); return
    data = make_excel(records)
    fname = f"participants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await q.message.reply_document(
        document=io.BytesIO(data), filename=fname,
        caption="📊 Список участников воркшопа"
    )

async def cmd_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Быстрый экспорт командой /export"""
    if not is_admin(update.effective_user.id): return
    records = db_all()
    if not records:
        await update.message.reply_text("Нет данных."); return
    if not HAS_OPENPYXL:
        await update.message.reply_text("⚠️ Установи openpyxl: pip install openpyxl"); return
    data = make_excel(records)
    fname = f"participants_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await update.message.reply_document(
        document=io.BytesIO(data), filename=fname,
        caption=f"📊 Участники воркшопа — {datetime.now().strftime('%d.%m.%Y')}"
    )

async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    st = db_stats()
    await update.message.reply_text(
        f"📊 *Статистика*\n\n"
        f"👥 Заявок: *{st['total']}*\n"
        f"✅ Оплачено: *{st['paid']}*\n"
        f"⏳ Ожидают: *{st['pending']}*\n"
        f"💰 Собрано: *{st['amount']} ₽*\n\n"
        f"📅 {WORKSHOP_DATE_STR}",
        parse_mode=ParseMode.MARKDOWN
    )

async def cmd_admin_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    await update.message.reply_text(
        "🛠 *Команды администратора*\n\n"
        "/participants — список участников\n"
        "/export — скачать Excel\n"
        "/stats — статистика\n"
        "/broadcast — рассылка сообщения\n"
        "/confirm\\_ID\\_СУММА — подтвердить оплату\n"
        "/reply\\_ID текст — ответить участнице\n\n"
        "_Примеры:_\n"
        "`/confirm_123456789_10000`\n"
        "`/reply_123456789 Добрый день!`",
        parse_mode=ParseMode.MARKDOWN
    )

# ═══════════════════════════════════════════════════════════
#  РАССЫЛКА
# ═══════════════════════════════════════════════════════════
async def cmd_broadcast(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    if not is_admin(update.effective_user.id): return ConversationHandler.END
    kb = [
        [InlineKeyboardButton("✅ Только оплатившим",           callback_data="bc_paid")],
        [InlineKeyboardButton("👥 Всем зарегистрированным",     callback_data="bc_all")],
        [InlineKeyboardButton("❌ Отмена",                      callback_data="bc_cancel")],
    ]
    await update.message.reply_text("📢 *Рассылка*\n\nКому отправить?",
                                    parse_mode=ParseMode.MARKDOWN,
                                    reply_markup=InlineKeyboardMarkup(kb))
    return S_BROADCAST_TARGET

async def cb_bc_target(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query; await q.answer()
    if q.data == "bc_cancel":
        await q.edit_message_text("Рассылка отменена.")
        return ConversationHandler.END
    ctx.user_data["bc_target"] = q.data
    label = "оплатившим" if q.data == "bc_paid" else "всем зарегистрированным"
    await q.edit_message_text(
        f"📝 Напишите текст рассылки *{label}*.\n\n"
        f"Можно прикрепить фото с подписью.\n/cancel — отмена.",
        parse_mode=ParseMode.MARKDOWN
    )
    return S_BROADCAST_TEXT

async def rx_broadcast_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    target = ctx.user_data.get("bc_target", "bc_all")
    ids = db_paid_ids() if target == "bc_paid" else db_all_ids()
    if not ids:
        await update.message.reply_text("⚠️ Нет получателей.")
        return ConversationHandler.END
    text  = update.message.text or update.message.caption or ""
    photo = update.message.photo[-1].file_id if update.message.photo else None
    await update.message.reply_text(f"⏳ Отправляю {len(ids)} получателям...")
    ok = fail = 0
    for uid in ids:
        try:
            if photo:
                await ctx.bot.send_photo(chat_id=uid, photo=photo,
                                         caption=text, parse_mode=ParseMode.MARKDOWN)
            else:
                await ctx.bot.send_message(chat_id=uid, text=text,
                                           parse_mode=ParseMode.MARKDOWN)
            ok += 1
            await asyncio.sleep(0.05)
        except Exception as e:
            logging.warning(f"Broadcast {uid}: {e}"); fail += 1
    await update.message.reply_text(f"✅ Готово!\n📬 Доставлено: {ok}\n❌ Ошибок: {fail}")
    return ConversationHandler.END

async def cmd_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Отменено. /start — начать заново.",
                                    reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# ═══════════════════════════════════════════════════════════
#  АВТОВЫДАЧА ДОСТУПА (вызывается из вебхука Тильды)
# ═══════════════════════════════════════════════════════════
async def grant_access(bot, rec: dict, amount: str):
    """Подтверждает оплату участнику и выдаёт доступ:
    сообщение-подтверждение + персональная одноразовая ссылка в канал."""
    user_id   = rec["user_id"]
    full_name = rec.get("full_name") or "Участник"

    # 1. Подтверждение
    await bot.send_message(
        chat_id=user_id,
        text=T_PAYMENT_CONFIRMED.format(full_name=full_name, date=WORKSHOP_DATE_STR),
        parse_mode=ParseMode.MARKDOWN,
    )

    # 2. Персональная ссылка-приглашение в закрытый канал
    await send_channel_invite(bot, user_id, full_name)

# ═══════════════════════════════════════════════════════════
#  HTTP-ЭНДПОИНТ ДЛЯ ВЕБХУКА ТИЛЬДЫ
# ═══════════════════════════════════════════════════════════
def _extract_payment(data: dict) -> tuple[str, str, str]:
    """Достаёт email, телефон и сумму из payload Тильды.
    Тильда шлёт поля формы плоско (Name/Email/Phone), а данные оплаты —
    в поле `payment` (JSON-строка с amount/orderid/products)."""
    email = (data.get("email") or data.get("Email") or data.get("E-mail") or "").strip().lower()
    phone = str(data.get("phone") or data.get("Phone") or data.get("tel") or "").strip()
    amount = str(data.get("amount") or data.get("payment_amount") or "")

    pay = data.get("payment")
    if pay:
        try:
            pay_obj = pay if isinstance(pay, dict) else json.loads(pay)
            amount = str(pay_obj.get("amount") or amount)
            if not email:
                email = (pay_obj.get("email") or "").strip().lower()
            if not phone:
                phone = str(pay_obj.get("phone") or "").strip()
        except (ValueError, TypeError):
            pass

    # Резервный парсинг: у Тильды поля могут приходить с кастомными именами
    # (например, Телефон, phone_123, custom_email и т.д.).
    if not email or not phone:
        for k, v in data.items():
            key = str(k).strip().lower()
            val = str(v or "").strip()
            if not val:
                continue
            if (not email) and ("@" in val and "." in val.split("@")[-1]):
                if any(token in key for token in ("email", "e-mail", "mail", "почт")):
                    email = val.lower()
            if not phone:
                if any(token in key for token in ("phone", "tel", "тел", "моб")):
                    if phone_tail10(val):
                        phone = val

    # Крайний fallback: если ключи странные, ищем похожие значения без привязки к имени поля
    if not email or not phone:
        for v in data.values():
            val = str(v or "").strip()
            if not val:
                continue
            if (not email) and ("@" in val and "." in val.split("@")[-1]):
                email = val.lower()
            if (not phone) and phone_tail10(val):
                # Игнорируем суммы/ID: номер телефона обычно >=10 цифр.
                if len("".join(ch for ch in val if ch.isdigit())) >= 10:
                    phone = val

    if not amount:
        amount = str(WORKSHOP_PRICE)
    return email, phone, amount

_bg_tasks: set = set()

def _spawn(coro):
    """Запускает фоновую корутину и держит ссылку, чтобы её не собрал GC."""
    t = asyncio.create_task(coro)
    _bg_tasks.add(t)
    t.add_done_callback(_bg_tasks.discard)

async def _notify_admins(bot, text: str):
    for aid in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=aid, text=text, parse_mode=ParseMode.MARKDOWN)
        except Exception:
            pass

async def _grant_and_notify(bot, rec: dict, email: str, phone: str, amount: str):
    """Фоновая выдача доступа — чтобы вебхук успел ответить Тильде за 5 сек."""
    try:
        await grant_access(bot, rec, amount)
    except Exception as e:
        logging.error(f"grant_access error: {e}")
    await _notify_admins(
        bot,
        (f"✅ Авто-подтверждение оплаты с Тильды\n"
         f"👤 {rec.get('full_name')}\n📧 {email or '—'}  📱 {phone or '—'}\n💰 {amount} ₽\n"
         f"🆔 `{rec['user_id']}` — доступ выдан.")
    )

async def handle_tilda_webhook(request: web.Request) -> web.Response:
    app = request.app["bot_app"]
    bot = app.bot

    # Проверка секрета (из query ?secret=... или заголовка)
    secret = request.query.get("secret") or request.headers.get("X-Tilda-Secret", "")
    if TILDA_WEBHOOK_SECRET and secret != TILDA_WEBHOOK_SECRET:
        logging.warning("Tilda webhook: неверный секрет")
        return web.Response(status=403, text="forbidden")

    # Парсим тело: form-urlencoded (по умолчанию у Тильды) или JSON
    data: dict = {}
    try:
        if "application/json" in (request.content_type or ""):
            data = await request.json()
        else:
            data = dict(await request.post())
    except Exception as e:
        logging.error(f"Tilda webhook parse error: {e}")

    # Тестовый пинг Тильды при сохранении вебхука
    if not data or "test" in data:
        return web.Response(text="ok")

    email, phone, amount = _extract_payment(data)
    logging.info(
        f"Tilda webhook: email={email or '-'}, phone_raw={phone or '-'}, "
        f"phone_norm={norm_phone(phone) or '-'}, amount={amount}"
    )

    if not email and not phone:
        for aid in ADMIN_IDS:
            try:
                await bot.send_message(
                    chat_id=aid,
                    text=f"⚠️ Пришла оплата с Тильды без email и телефона — выдайте доступ вручную.\n\n`{data}`"[:4000],
                    parse_mode=ParseMode.MARKDOWN,
                )
            except Exception:
                pass
        return web.Response(text="ok")

    rec = db_find_by_contact(email, phone)
    if not rec:
        # Человек оплатил на сайте раньше, чем зарегистрировался в боте.
        # Сохраняем оплату — доступ выдастся автоматически при регистрации.
        db_add_pending(email, phone, amount)
        _spawn(_notify_admins(
            bot,
            (f"💸 Оплата с сайта ({amount} ₽), но участник ещё не в боте.\n"
             f"📧 {email or '—'}  📱 {phone or '—'}\n"
             f"Оплата сохранена — доступ выдастся автоматически, когда человек "
             f"зарегистрируется в боте с тем же email/телефоном.")
        ))
        return web.Response(text="ok")

    # Повторный вебхук Тильды по уже оплаченной заявке — не дублируем выдачу
    if rec.get("status") == "Оплачено":
        return web.Response(text="ok")

    # Участник найден → подтверждаем; саму выдачу делаем в фоне, чтобы
    # ответить Тильде за 5 секунд (иначе она шлёт повторы).
    db_confirm(rec["user_id"], amount)
    _spawn(_grant_and_notify(bot, rec, email, phone, amount))
    return web.Response(text="ok")

async def handle_health(request: web.Request) -> web.Response:
    return web.Response(text="bot ok")

# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════
def build_app() -> Application:
    app = Application.builder().token(BOT_TOKEN).build()

    reg_conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            S_CONSENT:    [CallbackQueryHandler(cb_show_docs,    pattern="^show_docs$"),
                           CallbackQueryHandler(cb_read_offer,   pattern="^read_offer$"),
                           CallbackQueryHandler(cb_read_privacy, pattern="^read_privacy$"),
                           CallbackQueryHandler(cb_consent_yes,  pattern="^consent_yes$"),
                           CallbackQueryHandler(cb_consent_no,   pattern="^consent_no$")],
            S_FULLNAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_fullname)],
            S_PHONE:      [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), rx_phone)],
            S_EMAIL:      [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_email)],
            S_PAYMENT:    [CallbackQueryHandler(cb_payment_done, pattern="^payment_done$")],
            S_SCREENSHOT: [MessageHandler(filters.PHOTO | filters.Document.IMAGE |
                                          (filters.TEXT & ~filters.COMMAND), rx_screenshot)],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    support_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^💬 Поддержка$"), start_support)],
        states={S_SUPPORT: [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_support),
                             CallbackQueryHandler(cb_support_cancel, pattern="^support_cancel$")]},
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
    )

    broadcast_conv = ConversationHandler(
        entry_points=[CommandHandler("broadcast", cmd_broadcast)],
        states={
            S_BROADCAST_TARGET: [CallbackQueryHandler(cb_bc_target, pattern="^bc_")],
            S_BROADCAST_TEXT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_broadcast_text),
                                  MessageHandler(filters.PHOTO, rx_broadcast_text)],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
    )

    edit_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✏️ Изменить данные$"), menu_edit_data)],
        states={
            S_FULLNAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_fullname)],
            S_PHONE:      [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), rx_phone)],
            S_EMAIL:      [MessageHandler(filters.TEXT & ~filters.COMMAND, rx_email)],
            S_PAYMENT:    [CallbackQueryHandler(cb_payment_done, pattern="^payment_done$")],
            S_SCREENSHOT: [MessageHandler(filters.PHOTO | filters.Document.IMAGE |
                                          (filters.TEXT & ~filters.COMMAND), rx_screenshot)],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
    )

    app.add_handler(reg_conv)
    app.add_handler(edit_conv)
    app.add_handler(support_conv)
    app.add_handler(broadcast_conv)

    app.add_handler(MessageHandler(filters.Regex("^📋 Моя регистрация$"), menu_my_reg))
    app.add_handler(MessageHandler(filters.Regex("^📞 Связаться с организатором$"), menu_contact))
    app.add_handler(MessageHandler(filters.Regex("^💳 Повторить оплату$"), menu_resend_qr))
    app.add_handler(MessageHandler(filters.Regex("^🔄 Проверить оплату$"), menu_check_payment))
    app.add_handler(MessageHandler(filters.Regex("^🔗 Ссылка в канал$"), menu_send_channel_link))
    app.add_handler(MessageHandler(filters.Regex("^📄 Оферта$"),          menu_offer))
    app.add_handler(MessageHandler(filters.Regex("^🔒 Политика данных$"), menu_privacy))
    app.add_handler(CallbackQueryHandler(cb_go_support, pattern="^go_support$"))

    app.add_handler(CommandHandler("participants", cmd_participants))
    app.add_handler(CommandHandler("export",       cmd_export))
    app.add_handler(CommandHandler("stats",        cmd_stats))
    app.add_handler(CommandHandler("admin",        cmd_admin_help))
    app.add_handler(CallbackQueryHandler(cb_export_excel, pattern="^export_excel$"))
    app.add_handler(MessageHandler(
        filters.Regex(r"^/confirm_\d+_\d+") & filters.ChatType.PRIVATE, cmd_confirm))
    app.add_handler(MessageHandler(
        filters.Regex(r"^/reply_\d+") & filters.ChatType.PRIVATE, cmd_reply))

    # Контроль доступа в закрытый канал
    app.add_handler(ChatJoinRequestHandler(on_join_request))
    app.add_handler(ChatMemberHandler(on_chat_member_update, ChatMemberHandler.CHAT_MEMBER))

    return app

def make_web_app(bot_app: Application) -> web.Application:
    web_app = web.Application()
    web_app["bot_app"] = bot_app
    web_app.router.add_post("/tilda-webhook", handle_tilda_webhook)
    web_app.router.add_get("/tilda-webhook", handle_tilda_webhook)   # для теста Тильды
    web_app.router.add_get("/", handle_health)                       # health-check Railway
    return web_app

async def run():
    init_db()
    app = build_app()

    await app.initialize()
    await app.start()
    await app.updater.start_polling(drop_pending_updates=True)

    runner = web.AppRunner(make_web_app(app))
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()

    print(f"🌸 Бот запущен (polling) + HTTP-сервер на :{PORT}")
    print(f"   Вебхук Тильды: POST /tilda-webhook  | БД: {DB_PATH}")

    stop = asyncio.Event()
    try:
        await stop.wait()
    finally:
        await app.updater.stop()
        await app.stop()
        await app.shutdown()
        await runner.cleanup()

def main():
    logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
    try:
        asyncio.run(run())
    except (KeyboardInterrupt, SystemExit):
        print("Остановлено.")

if __name__ == "__main__":
    main()
